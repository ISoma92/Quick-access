import openpyxl

wb = openpyxl.load_workbook('MOCK_DATA.xlsx')
sheet = wb.active

iter_a = iter(sheet.iter_rows(min_col=1, max_col=1, min_row=2, values_only=True))
iter_b = iter(sheet.iter_rows(min_col=2, max_col=2, min_row=2, values_only=True))
iter_c = iter(sheet.iter_rows(min_col=3, max_col=3, min_row=2, values_only=True))
iter_d = iter(sheet.iter_rows(min_col=4, max_col=4, min_row=2, values_only=True))
iter_e = iter(sheet.iter_rows(min_col=5, max_col=5, min_row=2, values_only=True))
iter_f = iter(sheet.iter_rows(min_col=6, max_col=6, min_row=2, values_only=True))
iter_g = iter(sheet.iter_rows(min_col=7, max_col=7, min_row=2, values_only=True))

for row_a, row_b, row_c, row_d, row_e, row_f, row_g in zip(iter_a, iter_b, iter_c, iter_d, iter_e, iter_f, iter_g):
    if row_a[0] is not None and row_b[0] is not None and row_c[0] is not None and row_d[0] is not None and row_e[0] is not None and row_f[0] is not None and row_g[0] is not None:
        a = str(row_a[0])
        b = str(row_b[0])
        c = str(row_c[0])
        d = str(row_d[0])
        e = str(row_e[0])
        f = str(row_f[0])
        g = str(row_g[0])
        sql = f"INSERT INTO [HomeDb].[dbo].[Excel2SQLwithTXT] (id, first_name, last_name, email, gender, ip_address, country_code) VALUES ('{a}', '{b}', '{c}', '{d}', '{e}', '{f}', '{g}')"
        print(sql)
