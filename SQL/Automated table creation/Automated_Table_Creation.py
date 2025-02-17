import openpyxl
import pytds
from datetime import datetime

IndexDict = {'ID': 'numeric(18, 0)',
             'first_name': 'varchar(50)',
             'last_name': 'varchar(50)',
             'email': 'VARCHAR(255) UNIQUE NOT NULL',
             'gender': 'VARCHAR(10) CHECK (gender IN (\'Male\', \'Female\', \'Other\'))',
             'ip_address': 'varchar(50)',
             'country_code': 'VARCHAR(3)',
             'zipcode': 'NVARCHAR(20)',
             'dob': 'DATE',
             'phone_number': 'NVARCHAR(20)',
             'is_active': 'BIT DEFAULT 1',
             'price': 'DECIMAL(10,2)',
             'quantity': 'INT DEFAULT 0',
             "notes": "TEXT",
             "ordernumber": "INT"}

wb = openpyxl.load_workbook('MOCK_DATA2.xlsx')
ws = wb.active

conn = pytds.connect(
    server='DESKTOP-S07K9GM',
    database='HomeDb',
    user='Szofi',
    # password='your_password',  # Replace with your actual password if needed
    auth=pytds.login.SspiAuth(),
    autocommit=True
)
cursor = conn.cursor()

for rows in ws.iter_rows(min_row=1, max_row=1, values_only=True):
    columns = rows

# Table name and index
table_name = input("Enter the table name: ")

create_table_query = f"CREATE TABLE [{table_name}] (\n"
for column in columns:
    column_name = column.strip()
    data_type = IndexDict.get(column_name, 'VARCHAR(255)')  # Default to VARCHAR(255) if not found in IndexDict
    create_table_query += f"    [{column_name}] {data_type},\n"
create_table_query = create_table_query.rstrip(",\n") + "\n);"

# Create table
cursor.execute(create_table_query)
print(f"Table '{table_name}' created successfully with columns: {', '.join(columns)}")

# Insert data into the table
for row in ws.iter_rows(min_row=2, values_only=True):
    formatted_row = []
    for value, column in zip(row, columns):
        if IndexDict.get(column) == 'DATE' and isinstance(value, str):
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
                try:
                    value = datetime.strptime(value, fmt).date()
                    break
                except ValueError:
                    continue
            else:
                value = None  # Handle invalid date format
        formatted_row.append(value)
    insert_data_query = f"INSERT INTO [{table_name}] ({', '.join([f'[{col}]' for col in columns])}) VALUES ({', '.join(['%s'] * len(formatted_row))})"
    cursor.execute(insert_data_query, formatted_row)
    print(f"Data '{formatted_row}' inserted successfully.")

conn.close()