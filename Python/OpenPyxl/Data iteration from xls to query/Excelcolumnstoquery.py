import openpyxl

wb=openpyxl.load_workbook('sampledataworkorders.xlsx')
sheet=wb.active     

iter_b = iter(sheet.iter_rows(min_col=2, max_col=2, min_row=2, values_only=True))
iter_m = iter(sheet.iter_rows(min_col=13, max_col=13, min_row=2, values_only=True))


for row_b, row_m in zip(iter_b, iter_m):
    if row_b != None and row_m != None:
        b = "".join(str(ele) for ele in row_b)
        m = "".join(str(ele) for ele in row_m)
        sqlquery= f"District = '{b}' and Payment = '{m}' OR"
        print(sqlquery)